#!/usr/bin/env python3
"""
STH Prices Server v3.0
Servidor Flask para extraer precios de metales desde Transamine
Puerto: 8080

CORREGIDO: Extrae datos de spans, no de tablas HTML
"""

import os
import math
import time
import tempfile
import re
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

import requests
from bs4 import BeautifulSoup

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)
CORS(app)

TRANSAMINE_URL = "https://www.transamine.com/price-and-review.html"
REQUEST_TIMEOUT = 30
PAUSE_BETWEEN_MONTHS = 1

# Feeds JSON públicos de LBMA (fuente de RESPALDO para Au/Ag).
# Solo se consultan cuando Transamine devuelve nulls en oro/plata.
# Cada registro: {"d": "YYYY-MM-DD", "v": [USD, GBP, EUR]}
LBMA_FEEDS = {
    'oro_am': 'https://prices.lbma.org.uk/json/gold_am.json',
    'oro_pm': 'https://prices.lbma.org.uk/json/gold_pm.json',
    'plata': 'https://prices.lbma.org.uk/json/silver.json',
}
LBMA_TIMEOUT = 30

# Tablas publicas de Westmetall (fuente de RESPALDO para metales BASE).
# Republican el LME Cash-Settlement oficial diario — validado identico a
# Transamine (Pb 10-jul-2026: 1851 en ambos). Cubren el anio en curso.
# Solo se consultan cuando Transamine deja nulls en metales base.
WESTMETALL_URL = "https://www.westmetall.com/en/markdaten.php?action=table&field={field}"
WESTMETALL_FIELDS = {
    'cobre': 'LME_Cu_cash',
    'plomo': 'LME_Pb_cash',
    'zinc': 'LME_Zn_cash',
    'niquel': 'LME_Ni_cash',
    'estano': 'LME_Sn_cash',
}
WESTMETALL_TIMEOUT = 30
MONTHS_EN = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5,
             'June': 6, 'July': 7, 'August': 8, 'September': 9,
             'October': 10, 'November': 11, 'December': 12}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Connection': 'keep-alive',
}


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ['nan', 'none', 'null', 'n/a', '-', '']:
            return None
    return value


def clean_dict(d):
    if isinstance(d, dict):
        return {k: clean_dict(v) for k, v in d.items()}
    elif isinstance(d, list):
        return [clean_dict(item) for item in d]
    else:
        return clean_value(d)


def parse_price(price_str):
    if not price_str:
        return None
    try:
        cleaned = price_str.replace('$', '').replace(',', '').replace(' ', '').strip()
        if cleaned.lower() in ['n/a', '-', '', 'none', 'nan']:
            return None
        return float(cleaned)
    except (ValueError, AttributeError):
        return None


def get_months_in_range(start_date, end_date):
    months = []
    current = start_date.replace(day=1)
    while current <= end_date:
        months.append(current.strftime('%Y-%m'))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return months


def extract_month_prices(session, year_month):
    """
    Extrae precios de un mes desde Transamine.
    La página usa spans con clases text_price, no tablas HTML.
    """
    url = f"{TRANSAMINE_URL}?choix_date={year_month}"
    month_data = {}
    averages = {}
    
    try:
        response = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'lxml')
        
        # Buscar el contenedor principal
        table_listing = soup.find('div', id='table_listing')
        
        if not table_listing:
            print(f"  No se encontró table_listing para {year_month}")
            return month_data, averages
        
        # Encontrar todas las secciones (h4 seguido de div)
        current_metal = None
        
        # Iterar por los elementos hijos
        for element in table_listing.children:
            if element.name == 'h4':
                # Es un encabezado de metal
                metal_text = element.get_text().strip().lower()
                if 'copper' in metal_text:
                    current_metal = 'cobre'
                elif 'lead' in metal_text:
                    current_metal = 'plomo'
                elif 'zinc' in metal_text:
                    current_metal = 'zinc'
                elif 'nickel' in metal_text:
                    current_metal = 'niquel'
                elif 'gold' in metal_text:
                    current_metal = 'oro'
                elif 'silver' in metal_text:
                    current_metal = 'plata'
                elif 'tin' in metal_text:
                    current_metal = 'estano'
                elif 'cobalt' in metal_text:
                    current_metal = 'cobalto'
                else:
                    current_metal = None
                    
            elif element.name == 'div' and current_metal:
                # Es el contenedor de precios para el metal actual
                
                # Extraer promedio
                avg_span = element.find('span', class_='average_price')
                if avg_span:
                    avg_values = avg_span.find_all('span')
                    if current_metal == 'oro' and len(avg_values) >= 3:
                        # Oro tiene AM, MEAN, PM - tomamos PM (índice 2)
                        averages[f'{current_metal}_pm'] = parse_price(avg_values[2].get_text())
                        averages[f'{current_metal}_am'] = parse_price(avg_values[0].get_text())
                        averages[f'{current_metal}_mean'] = parse_price(avg_values[1].get_text())
                    elif avg_values:
                        averages[current_metal] = parse_price(avg_values[0].get_text())
                
                # Extraer precios diarios
                price_spans = element.find_all('span', class_='text_price')
                
                for price_span in price_spans:
                    strong = price_span.find('strong')
                    if not strong:
                        continue
                    
                    date_text = strong.get_text().strip()
                    
                    # Validar formato de fecha YYYY-MM-DD
                    if not re.match(r'\d{4}-\d{2}-\d{2}', date_text):
                        continue

                    # Descartar fechas de OTRO mes: cuando el feed de un metal
                    # esta roto, Transamine muestra el bloque del mes anterior
                    # (ej. Gold con fechas de junio en la pagina de julio) y esas
                    # filas parciales pisarian las filas completas del mes
                    # correcto en all_prices.update()
                    if not date_text.startswith(year_month):
                        continue
                    
                    # Inicializar fecha si no existe
                    if date_text not in month_data:
                        month_data[date_text] = {
                            'fecha': date_text,
                            'oro_pm': None,
                            'oro_am': None,
                            'plata': None,
                            'cobre': None,
                            'plomo': None,
                            'zinc': None,
                            'niquel': None,
                            'estano': None
                        }
                    
                    # Extraer valores
                    value_spans = price_span.find_all('span')
                    
                    if current_metal == 'oro' and len(value_spans) >= 3:
                        # Oro: AM, MEAN, PM
                        month_data[date_text]['oro_am'] = parse_price(value_spans[0].get_text())
                        month_data[date_text]['oro_pm'] = parse_price(value_spans[2].get_text())
                    elif value_spans:
                        price_value = parse_price(value_spans[0].get_text())
                        month_data[date_text][current_metal] = price_value
        
        print(f"  {year_month}: {len(month_data)} días extraídos")
        if averages:
            print(f"  Promedios: {averages}")
        
        return month_data, averages
        
    except requests.exceptions.Timeout:
        print(f"  Timeout para {year_month}")
        return month_data, averages
    except requests.exceptions.RequestException as e:
        print(f"  Error de conexión para {year_month}: {e}")
        return month_data, averages
    except Exception as e:
        print(f"  Error extrayendo precios para {year_month}: {e}")
        import traceback
        traceback.print_exc()
        return month_data, averages


def fetch_lbma_backup(start_str, end_str):
    """
    Descarga los fixes LBMA (Au AM/PM y Ag, en USD) para el rango de fechas.
    Devuelve dict {fecha: {'oro_am': x, 'oro_pm': y, 'plata': z}}.
    Es fuente de RESPALDO: solo se llama si Transamine trae nulls.
    Si un feed falla, se omite ese campo sin tumbar la respuesta.
    """
    backup = {}
    for field, url in LBMA_FEEDS.items():
        try:
            resp = requests.get(url, headers=HEADERS, timeout=LBMA_TIMEOUT)
            resp.raise_for_status()
            for record in resp.json():
                d = record.get('d')
                if not d or d < start_str or d > end_str:
                    continue
                v = record.get('v') or []
                usd = v[0] if len(v) >= 1 else None
                if usd in (None, 0):
                    continue
                backup.setdefault(d, {})[field] = float(usd)
            print(f"  LBMA respaldo: {field} OK")
        except Exception as e:
            print(f"  LBMA respaldo: error en {field}: {e}")
    return backup


def fill_missing_precious(results, averages_by_month, start_str, end_str):
    """
    Rellena con LBMA únicamente los campos oro_am, oro_pm y plata que
    Transamine haya dejado en None. No sobreescribe valores existentes.
    Recalcula promedios mensuales de preciosos si venían vacíos.
    Devuelve la fuente efectiva de preciosos: 'transamine', 'lbma_respaldo' o 'mixto'.
    """
    precious_fields = ['oro_am', 'oro_pm', 'plata']

    existing = {row.get('fecha') for row in results}
    missing_nulls = any(
        row.get(f) is None for row in results for f in precious_fields
    )
    # Días hábiles (L-V) del rango que no aparecieron en Transamine
    missing_days = False
    day = datetime.strptime(start_str, '%Y-%m-%d')
    end_dt = datetime.strptime(end_str, '%Y-%m-%d')
    while day <= end_dt:
        ds = day.strftime('%Y-%m-%d')
        if day.weekday() < 5 and ds not in existing:
            missing_days = True
            break
        day += timedelta(days=1)

    if not (missing_nulls or missing_days):
        return 'transamine'

    backup = fetch_lbma_backup(start_str, end_str)
    if not backup:
        return 'transamine'

    filled_any = False
    kept_any = False
    for row in results:
        fecha = row.get('fecha')
        for f in precious_fields:
            if row.get(f) is not None:
                kept_any = True
            elif fecha in backup and backup[fecha].get(f) is not None:
                row[f] = backup[fecha][f]
                filled_any = True

    # Crear fechas que Transamine no publicó pero LBMA sí tiene
    existing_dates = {row.get('fecha') for row in results}
    for fecha in sorted(backup.keys()):
        if fecha in existing_dates:
            continue
        new_row = {
            'fecha': fecha,
            'oro_pm': backup[fecha].get('oro_pm'),
            'oro_am': backup[fecha].get('oro_am'),
            'plata': backup[fecha].get('plata'),
            'cobre': None,
            'plomo': None,
            'zinc': None,
            'niquel': None,
            'estano': None
        }
        results.append(new_row)
        filled_any = True
    results.sort(key=lambda r: r.get('fecha') or '')

    if filled_any:
        # Recalcular promedios mensuales de preciosos desde los datos diarios
        # ya completos (Transamine puede publicar promedios corruptos cuando
        # su feed de preciosos está roto)
        for month, avgs in averages_by_month.items():
            for f in precious_fields:
                vals = [
                    row[f] for row in results
                    if row.get(f) is not None
                    and str(row.get('fecha', '')).startswith(month)
                ]
                if vals:
                    avgs[f] = round(sum(vals) / len(vals), 2)
            # oro_mean = promedio de AM y PM si ambos existen
            if avgs.get('oro_am') and avgs.get('oro_pm'):
                avgs['oro_mean'] = round((avgs['oro_am'] + avgs['oro_pm']) / 2, 2)

    if not filled_any:
        return 'transamine'
    return 'mixto' if kept_any else 'lbma_respaldo'


def fetch_westmetall_backup(start_str, end_str):
    """
    Descarga el LME Cash-Settlement diario desde las tablas publicas de
    Westmetall para los metales base. Devuelve {fecha: {metal: valor}}.
    Es fuente de RESPALDO: solo se llama si Transamine trae nulls en base.
    Si una tabla falla, se omite ese metal sin tumbar la respuesta.
    """
    backup = {}
    for metal, field in WESTMETALL_FIELDS.items():
        try:
            url = WESTMETALL_URL.format(field=field)
            resp = requests.get(url, headers=HEADERS, timeout=WESTMETALL_TIMEOUT)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, 'lxml')
            n = 0
            for tr in soup.find_all('tr'):
                cells = [td.get_text().strip() for td in tr.find_all('td')]
                if len(cells) < 2:
                    continue
                m = re.match(r'(\d{1,2})\.\s+([A-Za-z]+)\s+(\d{4})', cells[0])
                if not m or m.group(2) not in MONTHS_EN:
                    continue
                fecha = "%04d-%02d-%02d" % (
                    int(m.group(3)), MONTHS_EN[m.group(2)], int(m.group(1)))
                if fecha < start_str or fecha > end_str:
                    continue
                value = parse_price(cells[1])  # columna Cash-Settlement
                if value is None:
                    continue
                backup.setdefault(fecha, {})[metal] = value
                n += 1
            print(f"  Westmetall respaldo: {metal} OK ({n} fechas)")
        except Exception as e:
            print(f"  Westmetall respaldo: error en {metal}: {e}")
    return backup


def fill_missing_base(results, averages_by_month, start_str, end_str):
    """
    Rellena con Westmetall (LME Cash-Settlement) unicamente los campos de
    metales base (cobre, plomo, zinc, niquel, estano) que Transamine haya
    dejado en None. No sobreescribe valores existentes. Recalcula promedios
    mensuales de base si rellenaron datos.
    Devuelve la fuente efectiva: 'transamine', 'westmetall_respaldo' o 'mixto'.
    """
    base_fields = list(WESTMETALL_FIELDS.keys())

    existing = {row.get('fecha') for row in results}
    missing_nulls = any(
        row.get(f) is None for row in results for f in base_fields
    )
    # Dias habiles (L-V) del rango que no aparecieron en Transamine
    missing_days = False
    day = datetime.strptime(start_str, '%Y-%m-%d')
    end_dt = datetime.strptime(end_str, '%Y-%m-%d')
    while day <= end_dt:
        ds = day.strftime('%Y-%m-%d')
        if day.weekday() < 5 and ds not in existing:
            missing_days = True
            break
        day += timedelta(days=1)

    if not (missing_nulls or missing_days):
        return 'transamine'

    backup = fetch_westmetall_backup(start_str, end_str)
    if not backup:
        return 'transamine'

    filled_any = False
    kept_any = False
    for row in results:
        fecha = row.get('fecha')
        for f in base_fields:
            if row.get(f) is not None:
                kept_any = True
            elif fecha in backup and backup[fecha].get(f) is not None:
                row[f] = backup[fecha][f]
                filled_any = True

    # Crear fechas que Transamine no publico pero Westmetall si tiene
    existing_dates = {row.get('fecha') for row in results}
    for fecha in sorted(backup.keys()):
        if fecha in existing_dates:
            continue
        new_row = {
            'fecha': fecha,
            'oro_pm': None,
            'oro_am': None,
            'plata': None,
            'cobre': backup[fecha].get('cobre'),
            'plomo': backup[fecha].get('plomo'),
            'zinc': backup[fecha].get('zinc'),
            'niquel': backup[fecha].get('niquel'),
            'estano': backup[fecha].get('estano')
        }
        results.append(new_row)
        filled_any = True
    results.sort(key=lambda r: r.get('fecha') or '')

    if filled_any:
        # Recalcular promedios mensuales de base desde los datos diarios
        # ya completos
        for month, avgs in averages_by_month.items():
            for f in base_fields:
                vals = [
                    row[f] for row in results
                    if row.get(f) is not None
                    and str(row.get('fecha', '')).startswith(month)
                ]
                if vals:
                    avgs[f] = round(sum(vals) / len(vals), 2)

    if not filled_any:
        return 'transamine'
    return 'mixto' if kept_any else 'westmetall_respaldo'


@app.route('/')
def index():
    return jsonify({
        'status': 'ok',
        'message': 'STH Prices API',
        'version': '3.3.0',
        'endpoints': {
            '/': 'GET - Este mensaje',
            '/extract_prices': 'POST - Extraer precios (fecha_inicio, fecha_fin)',
            '/generate_excel': 'POST - Generar Excel con precios'
        }
    })


@app.route('/extract_prices', methods=['POST'])
def extract_prices():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify(clean_dict({'error': 'No se recibieron datos JSON'})), 400
        
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify(clean_dict({'error': 'Se requieren fecha_inicio y fecha_fin'})), 400
        
        try:
            start_date = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            end_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            return jsonify(clean_dict({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'})), 400
        
        if start_date > end_date:
            return jsonify(clean_dict({'error': 'fecha_inicio debe ser menor o igual a fecha_fin'})), 400
        
        months = get_months_in_range(start_date, end_date)
        print(f"Extrayendo {len(months)} mes(es): {', '.join(months)}")
        
        session = requests.Session()
        
        all_prices = {}
        all_averages = {}
        
        for i, month in enumerate(months):
            print(f"Descargando {month}...")
            month_prices, month_averages = extract_month_prices(session, month)
            all_prices.update(month_prices)
            all_averages[month] = month_averages
            
            if i < len(months) - 1:
                time.sleep(PAUSE_BETWEEN_MONTHS)
        
        session.close()
        
        # Filtrar fechas en el rango solicitado
        results = []
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            if date_str in all_prices:
                results.append(all_prices[date_str])
            current_date += timedelta(days=1)
        
        # Respaldo LBMA: rellena Au AM/PM y Ag solo donde Transamine dejó nulls
        fuente_preciosos = fill_missing_precious(
            results, all_averages,
            start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
        )

        # Respaldo Westmetall: rellena metales base (LME Cash-Settlement)
        # solo donde Transamine dejó nulls
        fuente_base = fill_missing_base(
            results, all_averages,
            start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
        )

        clean_results = clean_dict(results)

        print(f"Extracción completada: {len(clean_results)} fechas con datos "
              f"(preciosos: {fuente_preciosos}, base: {fuente_base})")

        return jsonify({
            'status': 'ok',
            'total_fechas': len(clean_results),
            'meses_descargados': len(months),
            'fuente_preciosos': fuente_preciosos,
            'fuente_base': fuente_base,
            'promedios': clean_dict(all_averages),
            'datos': clean_results
        })
        
    except Exception as e:
        print(f"Error en extract_prices: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(clean_dict({'error': str(e)})), 500


@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        data = request.get_json()
        
        if not data or 'datos' not in data:
            return jsonify(clean_dict({'error': 'Se requiere campo "datos" con precios'})), 400
        
        datos = data['datos']
        
        if not datos:
            return jsonify(clean_dict({'error': 'No hay datos para generar Excel'})), 400
        
        df = pd.DataFrame(datos)
        
        column_mapping = {
            'fecha': 'Fecha',
            'oro_pm': 'Oro PM (USD/oz)',
            'oro_am': 'Oro AM (USD/oz)',
            'plata': 'Plata (USD/oz)',
            'cobre': 'Cobre (USD/MT)',
            'plomo': 'Plomo (USD/MT)',
            'zinc': 'Zinc (USD/MT)',
            'niquel': 'Níquel (USD/MT)',
            'estano': 'Estaño (USD/MT)'
        }
        
        columns_to_use = [col for col in column_mapping.keys() if col in df.columns]
        df = df[columns_to_use]
        df = df.rename(columns=column_mapping)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Precios"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=clean_value(value))
                cell.border = thin_border
                
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    cell.alignment = Alignment(horizontal="center")
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"sth_prices_{timestamp}.xlsx"
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error en generate_excel: {e}")
        return jsonify(clean_dict({'error': str(e)})), 500


if __name__ == '__main__':
    print("=" * 50)
    print("STH Prices API Server v3.0")
    print("=" * 50)
    print("Puerto: 8080")
    print("URL: http://localhost:8080")
    print("=" * 50)
    print("CORREGIDO: Lee spans en lugar de tablas")
    print("=" * 50)
    
    app.run(host='0.0.0.0', port=8080, debug=True)
